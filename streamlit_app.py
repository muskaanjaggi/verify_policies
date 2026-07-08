"""
Policy PDF Verification — Streamlit App
=========================================
Upload an Excel file with policy details + PDF links, and this app will:
  1. Download each linked PDF
  2. Verify the policy number, start date, and end date against the PDF text
  3. Let you download a flagged/highlighted Excel file with the results

Deploy on Streamlit Community Cloud:
  1. Push this file + requirements.txt to a GitHub repo
  2. Go to share.streamlit.io, connect the repo, set main file to streamlit_app.py
  3. Done — no server setup needed

Required input columns in the uploaded Excel (case-sensitive):
  number, link, start_date, end_date
  (any other columns are preserved as-is in the output)
"""

import re
import io
import math
import time
import logging

import requests
import fitz  # pymupdf
import pandas as pd
import streamlit as st
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Optional OCR fallback for scanned/image-only PDFs
try:
    import pytesseract
    from pdf2image import convert_from_bytes
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

logging.basicConfig(level=logging.INFO)

REQUEST_CONNECT_TIMEOUT = 10
REQUEST_READ_TIMEOUT    = 45
DELAY_BETWEEN_REQUESTS  = 0.5
MAX_RETRIES             = 4
MIN_TEXT_CHARS_OK       = 20

REQUIRED_COLUMNS = ["number", "link", "start_date", "end_date"]


# ── Session with retries ──────────────────────────────────────────────────────

@st.cache_resource
def build_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=MAX_RETRIES,
        backoff_factor=1.5,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


# ── PDF fetching ──────────────────────────────────────────────────────────────

def fetch_pdf_text(session: requests.Session, url: str) -> tuple[str | None, str]:
    """Download PDF from url and return (text, error_detail)."""
    try:
        r = session.get(
            url,
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT),
        )
    except requests.exceptions.ConnectTimeout:
        return None, "Connection timed out"
    except requests.exceptions.ReadTimeout:
        return None, "Read timed out (file too large or slow response)"
    except requests.exceptions.SSLError as e:
        return None, f"SSL error: {e}"
    except requests.exceptions.ConnectionError as e:
        return None, f"Connection error: {e}"
    except requests.exceptions.RequestException as e:
        return None, f"Request failed: {e}"

    if r.status_code != 200:
        return None, f"HTTP {r.status_code} (expired/invalid link or access denied)"

    content = r.content
    if not content.startswith(b"%PDF"):
        snippet = content[:120].decode("utf-8", errors="replace")
        return None, f"Response was not a PDF (starts with: {snippet!r})"

    try:
        doc = fitz.open(stream=content, filetype="pdf")
    except Exception as e:
        return None, f"PyMuPDF could not open file: {e}"

    if doc.is_encrypted:
        if not doc.authenticate(""):
            return None, "PDF is password-protected and could not be opened"

    try:
        text_parts = []
        for page in doc:
            page_text = page.get_text()
            if len(page_text.strip()) < MIN_TEXT_CHARS_OK and OCR_AVAILABLE:
                page_text = ocr_page_fallback(content, page.number)
            text_parts.append(page_text)
        full_text = "\n".join(text_parts)
    except Exception as e:
        return None, f"Text extraction failed: {e}"

    if not full_text.strip():
        return None, "No extractable text (scanned PDF, OCR unavailable/failed)"

    return full_text, ""


def ocr_page_fallback(pdf_bytes: bytes, page_number: int) -> str:
    try:
        images = convert_from_bytes(
            pdf_bytes, first_page=page_number + 1, last_page=page_number + 1, dpi=300
        )
        if not images:
            return ""
        return pytesseract.image_to_string(images[0])
    except Exception as e:
        logging.info(f"OCR fallback failed on page {page_number}: {e}")
        return ""


# ── Policy number normalisation ───────────────────────────────────────────────

def normalise_policy_number(raw) -> str:
    if pd.isna(raw):
        return ""
    if isinstance(raw, float):
        if math.isnan(raw):
            return ""
        return str(int(raw))
    s = str(raw).strip()
    if re.fullmatch(r'[\d\.]+[eE][+\-]?\d+', s):
        try:
            return str(int(float(s)))
        except Exception:
            return s
    if re.fullmatch(r'\d+\.0', s):
        return s[:-2]
    return s


def policy_number_in_pdf(raw_excel_value, pdf_text: str) -> bool:
    if pd.isna(raw_excel_value):
        return True
    norm_val = normalise_policy_number(raw_excel_value)
    if not norm_val:
        return True
    pdf_norm = re.sub(r"\s+", " ", pdf_text).strip()
    if norm_val.lower() in pdf_norm.lower():
        return True
    digits_val = re.sub(r"[^0-9A-Za-z]", "", norm_val)
    digits_pdf = re.sub(r"[^0-9A-Za-z]", "", pdf_norm)
    if len(digits_val) >= 5 and digits_val.lower() in digits_pdf.lower():
        return True
    pure_digits = re.sub(r"\D", "", norm_val)
    if len(pure_digits) > 12:
        prefix = pure_digits[:12]
        pdf_digits = re.sub(r"\D", "", pdf_norm)
        if prefix in pdf_digits:
            return True
    return False


# ── Date extraction & matching ────────────────────────────────────────────────

def extract_all_dates_from_text(text: str) -> list[pd.Timestamp]:
    patterns = [
        r"\b(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{4})\b",
        r"\b(\d{4})[\/\-\.](\d{1,2})[\/\-\.](\d{1,2})\b",
        r"\b(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2})\b",
        r"\b(\d{1,2})[\s\-\/\.]+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-\/\.]+(\d{4})\b",
        r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-\/\.]+(\d{1,2})[\s,\-]+(\d{4})\b",
        r"\b(\d{1,2})[\s\-\/\.]+(January|February|March|April|May|June|July|August|September|October|November|December)[\s\-\/\.]+(\d{4})\b",
        r"\b(January|February|March|April|May|June|July|August|September|October|November|December)[\s\-\/\.]+(\d{1,2}),?[\s\-]+(\d{4})\b",
    ]
    month_map = {
        "jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
        "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12,
        "january":1,"february":2,"march":3,"april":4,"june":6,
        "july":7,"august":8,"september":9,"october":10,
        "november":11,"december":12,
    }
    found = []
    for pat in patterns:
        for m in re.finditer(pat, text, re.IGNORECASE):
            g = m.groups()
            try:
                if all(x.isdigit() for x in g):
                    a, b, c = int(g[0]), int(g[1]), int(g[2])
                    if c > 31:
                        ts = pd.Timestamp(year=a, month=b, day=c) if a > 31 \
                             else pd.Timestamp(year=c, month=b, day=a)
                    else:
                        year = 2000 + c if c < 50 else 1900 + c
                        ts = pd.Timestamp(year=year, month=b, day=a)
                else:
                    parts = [x for x in g if x]
                    nums  = [p for p in parts if p.isdigit()]
                    words = [p for p in parts if not p.isdigit()]
                    month = month_map.get(words[0].lower())
                    if not month:
                        continue
                    day  = int(nums[0]) if len(nums) >= 1 else None
                    year = int(nums[1]) if len(nums) >= 2 else int(nums[0])
                    if day is None or day > 31:
                        continue
                    ts = pd.Timestamp(year=year, month=month, day=day)
                if 1900 < ts.year < 2100:
                    found.append(ts)
            except Exception:
                continue
    return found


def date_matches(expected_dt, pdf_dates: list[pd.Timestamp]) -> bool:
    if pd.isna(expected_dt):
        return True
    try:
        exp = pd.Timestamp(expected_dt)
    except Exception:
        return True
    return any(
        d.year == exp.year and d.month == exp.month and d.day == exp.day
        for d in pdf_dates
    )


# ── Core verification loop ────────────────────────────────────────────────────

def run_verification(df: pd.DataFrame, progress_bar, status_text) -> pd.DataFrame:
    session = build_session()
    statuses, error_details = [], []
    total = len(df)

    for idx, row in df.iterrows():
        link = row.get("link")

        if pd.isna(link) or not str(link).strip().startswith("http"):
            statuses.append("")
            error_details.append("")
            progress_bar.progress((idx + 1) / total)
            continue

        status_text.text(f"Checking row {idx + 1} of {total} …")

        text, err_detail = fetch_pdf_text(session, str(link))
        if text is None:
            statuses.append("ERROR: Could not download/parse PDF")
            error_details.append(err_detail)
            progress_bar.progress((idx + 1) / total)
            time.sleep(DELAY_BETWEEN_REQUESTS * 3)
            continue

        issues = []
        pdf_dates = extract_all_dates_from_text(text)

        raw_num = row.get("number")
        if not pd.isna(raw_num) and str(raw_num).strip() not in ("", "nan"):
            if not policy_number_in_pdf(raw_num, text):
                display = normalise_policy_number(raw_num)
                issues.append(f"Policy number mismatch: expected '{display}'")

        start = row.get("start_date")
        if not date_matches(start, pdf_dates):
            label = pd.Timestamp(start).date() if not pd.isna(start) else start
            issues.append(f"Start date mismatch: expected '{label}'")

        end = row.get("end_date")
        if not date_matches(end, pdf_dates):
            label = pd.Timestamp(end).date() if not pd.isna(end) else end
            issues.append(f"End date mismatch: expected '{label}'")

        status = "PASS" if not issues else "FLAG: " + " | ".join(issues)
        statuses.append(status)
        error_details.append("")
        progress_bar.progress((idx + 1) / total)
        time.sleep(DELAY_BETWEEN_REQUESTS)

    df = df.copy()
    df["Verification_Status"] = statuses
    df["Download_Error_Detail"] = error_details
    return df


def build_highlighted_excel(df: pd.DataFrame) -> bytes:
    """Write df to an in-memory xlsx with colour-coded rows, return bytes."""
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    red_fill   = PatternFill("solid", fgColor="FFCCCC")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    err_fill   = PatternFill("solid", fgColor="FFD700")
    bold_red   = Font(bold=True, color="CC0000")
    bold_green = Font(bold=True, color="006400")

    header_row = [cell.value for cell in ws[1]]
    status_col = header_row.index("Verification_Status") + 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[status_col - 1]
        if not cell.value:
            continue
        if str(cell.value).startswith("FLAG"):
            for c in row:
                c.fill = red_fill
            cell.font = bold_red
        elif str(cell.value) == "PASS":
            cell.fill = green_fill
            cell.font = bold_green
        elif str(cell.value).startswith("ERROR"):
            cell.fill = err_fill

    ws.column_dimensions[ws.cell(1, status_col).column_letter].width = 80

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Streamlit UI ───────────────────────────────────────────────────────────────

st.set_page_config(page_title="Policy PDF Verifier", page_icon="📋", layout="centered")

st.title("📋 Policy PDF Verifier")
st.write(
    "Upload an Excel file with policy details and PDF links. "
    "Each PDF will be downloaded and checked against the policy number, "
    "start date, and end date in your sheet."
)

with st.expander("Required column format"):
    st.markdown(
        "Your Excel file must include these columns (case-sensitive):\n\n"
        "- **number** — policy number to verify\n"
        "- **link** — direct URL to the policy PDF\n"
        "- **start_date** — expected policy start date\n"
        "- **end_date** — expected policy end date\n\n"
        "Any other columns are kept as-is in the output file."
    )

if not OCR_AVAILABLE:
    st.info(
        "OCR fallback for scanned PDFs is not installed in this environment "
        "(pytesseract / pdf2image / poppler). Text-based PDFs will still work "
        "normally.",
        icon="ℹ️",
    )

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    try:
        df = pd.read_excel(uploaded_file, dtype={"number": object})
    except Exception as e:
        st.error(f"Could not read the Excel file: {e}")
        st.stop()

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        st.error(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Found columns: {', '.join(df.columns)}"
        )
        st.stop()

    st.success(f"Loaded {len(df)} rows.")
    st.dataframe(df.head(10), use_container_width=True)

    if st.button("Run verification", type="primary"):
        progress_bar = st.progress(0)
        status_text = st.empty()

        with st.spinner("Verifying policies …"):
            result_df = run_verification(df, progress_bar, status_text)

        status_text.empty()
        progress_bar.empty()

        checked = result_df["Verification_Status"][result_df["Verification_Status"] != ""]
        passed  = (checked == "PASS").sum()
        flagged = checked.str.startswith("FLAG").sum()
        errors  = checked.str.startswith("ERROR").sum()

        col1, col2, col3 = st.columns(3)
        col1.metric("✅ Passed", int(passed))
        col2.metric("🚩 Flagged", int(flagged))
        col3.metric("⚠️ Download errors", int(errors))

        st.subheader("Results preview")
        st.dataframe(result_df, use_container_width=True)

        excel_bytes = build_highlighted_excel(result_df)
        st.download_button(
            label="⬇️ Download flagged Excel file",
            data=excel_bytes,
            file_name="flagged.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
